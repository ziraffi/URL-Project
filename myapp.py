"""
Bulk URL Status Checker Web Application

This Flask web application provides functionality to check URL status codes and WHOIS information
for bulk URL processing. It supports uploading CSV/XLSX files or manual URL input, processes URLs
asynchronously, and provides downloadable results.

Features:
- File upload support (CSV, XLSX)
- Manual URL input
- Asynchronous URL processing with progress tracking
- WHOIS information retrieval
- Domain expiration status checking
- CSV result export
- Progress monitoring and cancellation

Routes:
- /: Main application page
- /input_section: Process uploaded files
- /process_manual_input: Process manually entered URLs
- /process_url_data: Main URL processing endpoint
- /progress: Progress monitoring endpoint
- /download/<filename>: Download processed results

Dependencies:
- Flask: Web framework
- pandas: Data processing
- openpyxl: Excel file handling
- aiohttp: Asynchronous HTTP requests
- whois: Domain information lookup
- tqdm: Progress bars
"""

# Configuration constants
REMAINING_ITERATIONS = 'remaining_iterations'
PROCESSED_URLS = 'processed_urls'
PROCESSING_STATUS = 'Processing'
STATUS_KEY = 'status'
TOTAL_URLS = 'total_urls'
PROGRESS_PERCENT = 'tryPercent'

# File system paths
TEMP_FILE_PATH = '/app/tmpf'

# Error messages
NO_FILE_UPLOADED = 'No file uploaded'
FILE_KEY = 'file'
XML_MIME_TYPE = 'application/xml'

# Progress tracking keys
ITERATION_TIMES = 'iteration_times'
URL_KEY = 'URL'
DOMAIN_INFO_KEY = 'domain_info'
CANCEL_FLAG = 'cancelFlag'
COLUMN_DATA = 'column_data'
COLUMN_NUMBER = 'column_number'
ROW_NUMBER = 'row_number'
SHEET_NUMBER = 'sheet_number'
SELECTED_COLUMN = 'selected_column'

# Built-in functions
ISINSTANCE_FUNC = isinstance
RANGE_FUNC = range
OPEN_FUNC = open

# Data processing keys
PROGRESS_INFO_OBJ = 'pInfo_obj'
DATA_KEY = 'data'
SELECTED_SHEET = 'selected_sheet'

# Utility functions
LEN_FUNC = len
ENUMERATE_FUNC = enumerate

# Domain analysis columns
EXPIRATION_DATE = 'Expiration Date'
RESPONSE_MESSAGE = 'Response Message'
STATUS_CODE = 'Status Code'

# Boolean constants
FALSE_VALUE = False
ERROR_MESSAGE_KEY = 'error_message'
POST_METHOD = 'POST'

# Exception handling
EXCEPTION_TYPE = Exception
ERROR_KEY = 'error'

# More boolean/constants
TRUE_VALUE = True
STRING_TYPE = str
NONE_VALUE = None
PRINT_FUNC = print

# Flask imports
from flask import Flask, render_template, request, jsonify, make_response, Response, send_from_directory

# Standard library imports
import json
import os
import random
import string
import uuid
from io import StringIO, BytesIO

# Third-party imports
import pandas as pd
import openpyxl
import logging

# Local imports
from domain_checker import process_urls_async, progress_info
# Flask app initialization
app = Flask(__name__)
user_sessions = {}

# Configure logging
logging.basicConfig(level=logging.DEBUG)
@app.route('/')
def home():
    return render_template('url_Project.html')

@app.route('/robots.txt')
def robots_txt():
    return send_from_directory('static', 'robots.txt')

@app.route('/set-cookie', methods=[POST_METHOD])
def set_cookie():
    USER_ID_KEY = 'user_id'
    COOKIE_CONSENT_KEY = 'cookie_consent'
    user_id = str(uuid.uuid4())
    user_sessions[user_id] = TRUE_VALUE
    cookie_data = {USER_ID_KEY: user_id, COOKIE_CONSENT_KEY: request.cookies.get(COOKIE_CONSENT_KEY)}
    storage_dir = '/app/storage'
    cookie_file_path = os.path.join(storage_dir, 'cookie_data.json')

    if not os.path.exists(cookie_file_path):
        with open(cookie_file_path, 'w') as file:
            json.dump([cookie_data], file, indent=4)
    else:
        with open(cookie_file_path, 'r+') as file:
            try:
                existing_data = json.load(file)
            except json.decoder.JSONDecodeError:
                existing_data = []
            existing_data.append(cookie_data)
            file.seek(0)
            json.dump(existing_data, file, indent=4)

    response = make_response()
    response.set_cookie(USER_ID_KEY, user_id, max_age=365*24*60*60)
    response.set_cookie(COOKIE_CONSENT_KEY, 'true', max_age=365*24*60*60)
    return response
@app.route('/guide.html')
def guide():
    return render_template('guide.html')

@app.route('/favicon.ico')
def favicon():
    return 'dummy', 200

@app.route('/disclaimer.html')
def disclaimer():
    return render_template('disclaimer.html')

@app.route('/privacyPolicy.html')
def privacy_policy():
    return render_template('PrivacyPolicy.html')
@app.route('/templates/sitemap.xml')
def sitemap_xml():
    sitemap_content = '''<?xml version="1.0" encoding="UTF-8"?>
    <urlset
      xmlns="http://www.sitemaps.org/schemas/sitemap/0.9"
      xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
      xsi:schemaLocation="http://www.sitemaps.org/schemas/sitemap/0.9
            http://www.sitemaps.org/schemas/sitemap/0.9/sitemap.xsd">

    <url>
      <loc>https://urlproject.azurewebsites.net/</loc>
      <lastmod>2024-04-05T15:09:03+00:00</lastmod>
      <priority>1.00</priority>
    </url>
    <url>
      <loc>https://urlproject.azurewebsites.net/templates/disclaimer.html</loc>
      <lastmod>2024-04-05T15:09:03+00:00</lastmod>
      <priority>0.80</priority>
    </url>
    <url>
      <loc>https://urlproject.azurewebsites.net/templates/privacyPolicy.html</loc>
      <lastmod>2024-04-05T15:09:03+00:00</lastmod>
      <priority>0.80</priority>
    </url>

    </urlset>
    '''
    return Response(sitemap_content, mimetype=XML_MIME_TYPE)
@app.route('/templates/ror.xml')
def ror_xml():
    ror_content = '''<?xml version="1.0" encoding="UTF-8"?>
    <rss version="2.0" xmlns:ror="http://rorweb.com/0.1/" >
    <channel>
      <title>ROR Sitemap for https://urlproject.azurewebsites.net/</title>
      <link>https://urlproject.azurewebsites.net/</link>

    <item>
         <link>https://urlproject.azurewebsites.net/</link>
         <title>Home Page</title>
         <description>Home Page</description>
         <ror:updatePeriod></ror:updatePeriod>
         <ror:sortOrder>0</ror:sortOrder>
         <ror:resourceOf>sitemap</ror:resourceOf>
    </item>
    <item>
         <link>https://urlproject.azurewebsites.net/templates/disclaimer.html</link>
         <title>Terms &amp; conditions</title>
         <description>Terms &amp; conditions</description>
         <ror:updatePeriod></ror:updatePeriod>
         <ror:sortOrder>1</ror:sortOrder>
         <ror:resourceOf>sitemap</ror:resourceOf>
    </item>
    <item>
         <link>https://urlproject.azurewebsites.net/templates/privacyPolicy.html</link>
         <title>Privacy Policy</title>
         <description>Privacy Policy</description>
         <ror:updatePeriod></ror:updatePeriod>
         <ror:sortOrder>1</ror:sortOrder>
         <ror:resourceOf>sitemap</ror:resourceOf>
    </item>
    </channel>
    </rss>
    '''
    return Response(ror_content, mimetype=XML_MIME_TYPE)

@app.route('/templates/sitemap.html')
def sitemap_html():
    return render_template('sitemap.html')

@app.route('/templates/<section_name>.html')
def template_section(section_name):
    return render_template(f"{section_name}.html"), 200
@app.route('/input_section', methods=[POST_METHOD])
def input_section():
    uploaded_file = request.files.get(FILE_KEY)
    selected_sheet = request.form.get(SELECTED_SHEET)
    selected_column = request.form.get(SELECTED_COLUMN)
    result = process_file_input(uploaded_file, selected_sheet, selected_column)
    if result:
        return jsonify(result)
    else:
        return jsonify({ERROR_MESSAGE_KEY: NO_FILE_UPLOADED})
@app.route('/process_manual_input', methods=[POST_METHOD])
def process_manual_input():
    try:
        request_data = request.get_json()
        manual_data = request_data['manual_data']
        selected_sheet = request_data[SELECTED_SHEET]
        selected_column = request_data[SELECTED_COLUMN]
        url_lines = manual_data.strip().split('\n')
        url_lines = [line.strip() for line in url_lines if line.strip()]
        csv_content = '\n'.join([f'"{line.replace(",", "")}"' for line in url_lines])
        df = pd.read_csv(StringIO(csv_content), header=NONE_VALUE, names=[selected_column])
        column_data = [{DATA_KEY: url, SHEET_NUMBER: 1, ROW_NUMBER: idx+2, COLUMN_NUMBER: 1}
                      for idx, url in enumerate(df[selected_column].tolist()) if pd.notna(url)]
        return jsonify({COLUMN_DATA: column_data, 'is_manual': TRUE_VALUE})
    except Exception as error:
        return jsonify({ERROR_MESSAGE_KEY: f"Error processing manual input: {str(error)}"}), 500
def process_file_input(uploaded_file, selected_sheet, selected_column):
    """
    Process uploaded file (CSV or XLSX) and extract URL data.

    Args:
        uploaded_file: Flask file object from request
        selected_sheet: Sheet name for XLSX files (ignored for CSV)
        selected_column: Column name containing URLs

    Returns:
        dict: Processing results with column data and metadata, or error message
    """
    IS_CSV_KEY = 'is_csv'
    SELECTED_FILE_KEY = 'selected_file'
    SHEET_COLUMNS_KEY = 'sheet_columns'
    sheet_name = selected_sheet
    file_obj = uploaded_file
    column_name = selected_column

    if not file_obj:
        return {ERROR_MESSAGE_KEY: NO_FILE_UPLOADED}

    try:
        file_content = file_obj.read()
        if file_obj.filename.endswith('.csv'):
            base_filename = file_obj.filename.split('.')[0]
            df = pd.read_csv(BytesIO(file_content))
            columns_list = list(df.columns)
            column_name = column_name or columns_list[0]
            column_data = [{DATA_KEY: url, SHEET_NUMBER: 1, ROW_NUMBER: idx+2, COLUMN_NUMBER: 1}
                          for idx, url in enumerate(df[column_name].tolist()) if pd.notna(url)]
            print('Current File name for CSV Test: ', base_filename)
            return {SHEET_COLUMNS_KEY: columns_list, COLUMN_DATA: column_data, SELECTED_SHEET: base_filename,
                   SELECTED_FILE_KEY: base_filename, IS_CSV_KEY: TRUE_VALUE}

        elif file_obj.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=TRUE_VALUE)
            sheet_names = workbook.sheetnames
            sheet_columns = {}
            column_data = []
            base_filename = file_obj.filename.split('.')[0]
            sheet_name = sheet_name or sheet_names[0]

            for sheet_idx, sheet_name_iter in enumerate(sheet_names, start=1):
                worksheet = workbook[sheet_name_iter]
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                columns_list = []

                if sheet_name_iter == sheet_name:
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row=1, column=col_idx)
                        col_name = cell.value
                        if col_name is None:
                            col_name = f"Column_{col_idx}"
                        columns_list.append(col_name)

                    if column_name is None:
                        column_name = columns_list[0]

                    if column_name not in columns_list:
                        return {ERROR_MESSAGE_KEY: f"Selected column '{column_name}' not found in sheet '{sheet_name}'."}

                    col_idx = columns_list.index(column_name) + 1
                    column_data = [{DATA_KEY: worksheet.cell(row=row_idx, column=col_idx).value,
                                  SHEET_NUMBER: sheet_idx, ROW_NUMBER: row_idx, COLUMN_NUMBER: col_idx}
                                 for row_idx in range(2, max_row + 1)
                                 if worksheet.cell(row=row_idx, column=col_idx).value]
                    column_name = columns_list[col_idx - 1]

                sheet_columns[sheet_name_iter] = columns_list

            return {SELECTED_FILE_KEY: base_filename, 'sheet_names': sheet_names,
                   SHEET_COLUMNS_KEY: sheet_columns, COLUMN_DATA: column_data,
                   'column_titles': columns_list, SELECTED_SHEET: sheet_name,
                   SELECTED_COLUMN: column_name, IS_CSV_KEY: FALSE_VALUE}
        else:
            return {ERROR_MESSAGE_KEY: 'Unsupported file format'}
    except Exception as error:
        return {ERROR_MESSAGE_KEY: f"Error processing file: {str(error)}"}
logging.basicConfig(level=logging.DEBUG)

def generate_unique_filename(base_filename):
    filename = base_filename
    dir_path = os.path.dirname(filename)
    base_name = os.path.basename(filename)
    if not os.path.exists(filename):
        return filename
    random_suffix = ''.join(random.choice(string.hexdigits) for _ in range(8))
    unique_filename = f"rk-{random_suffix}-{base_name}"
    return os.path.join(dir_path, unique_filename)
import os as E
from flask import request as C,jsonify as D
@app.route('/process_url_data', methods=[POST_METHOD])
async def process_url_data():
    """
    Main endpoint for processing URLs asynchronously.

    Expects JSON payload with:
    - clientUrlSet: Object containing url_list array
    - cancelFlag: Boolean to cancel processing

    Returns:
        JSON response with processing results or error messages
    """
    NO_DATA_MESSAGE = 'No data retrieved'
    MESSAGE_KEY = 'message'
    try:
        if not request.is_json:
            return jsonify({ERROR_KEY: 'Request content must be in JSON format'}), 400

        request_json = request.json
        client_url_set = request_json.get('clientUrlSet')
        cancel_flag = request_json.get(CANCEL_FLAG, FALSE_VALUE)

        if not client_url_set or 'url_list' not in client_url_set:
            return jsonify({ERROR_KEY: 'Invalid request: clientUrlSet with url_list required'}), 400

        url_list = client_url_set['url_list']
        if not isinstance(url_list, list) or not url_list:
            return jsonify({ERROR_KEY: 'URL list must be a non-empty array'}), 400

        if cancel_flag:
            return jsonify({MESSAGE_KEY: 'URL processing canceled'}), 200

        semaphore = asyncio.Semaphore(8)
        logging.info(f'Starting URL processing for {len(url_list)} URLs')
        domain_info_list = []
        yield_count = 0

        async for result in process_urls_async(url_list, semaphore, cancel_flag):
            yield_count += 1
            logging.info(f"Yield {yield_count}: Processing result: {result.keys()}, has_domain_info: {DOMAIN_INFO_KEY in result}")
            if DOMAIN_INFO_KEY in result:
                domain_info_list.append(result)
                logging.info(f"Added result to domain_info_list, total: {len(domain_info_list)}, domain_data type: {type(result[DOMAIN_INFO_KEY])}")
            if cancel_flag:
                logging.info("Cancel flag detected, breaking")
                break

        logging.info(f'URL processing completed, total yields: {yield_count}, domain_info_list length: {len(domain_info_list)}')

        if not domain_info_list:
            logging.warning('No domain information was retrieved')
            return jsonify({ERROR_KEY: NO_DATA_MESSAGE}), 404

        try:
            df_results = pd.DataFrame(domain_info_list)
            df_expanded = df_results[DOMAIN_INFO_KEY].apply(pd.Series)

            # Process status codes and response messages
            if STATUS_CODE in df_expanded.columns:
                df_expanded[STATUS_CODE] = df_expanded[STATUS_CODE].apply(join_array_values)
            if RESPONSE_MESSAGE in df_expanded.columns:
                df_expanded[RESPONSE_MESSAGE] = df_expanded[RESPONSE_MESSAGE].apply(join_array_values)

            # Process expiration dates
            if EXPIRATION_DATE in df_expanded.columns:
                try:
                    df_expanded[EXPIRATION_DATE] = pd.to_datetime(df_expanded[EXPIRATION_DATE], unit='ms', errors='coerce')
                    # Only apply timezone operations to valid datetime values
                    valid_dates_mask = df_expanded[EXPIRATION_DATE].notna()
                    if valid_dates_mask.any():
                        df_expanded.loc[valid_dates_mask, EXPIRATION_DATE] = df_expanded.loc[valid_dates_mask, EXPIRATION_DATE] \
                            .dt.tz_localize('UTC', errors='coerce') \
                            .dt.tz_convert('Asia/Kolkata', errors='coerce')
                        # Convert to string format for frontend safety
                        df_expanded[EXPIRATION_DATE] = df_expanded[EXPIRATION_DATE].astype(str)
                    else:
                        # Convert all NaT values to None/null for frontend
                        df_expanded[EXPIRATION_DATE] = None
                except Exception as date_error:
                    logging.warning(f"Error processing dates: {date_error}")
                    # Convert to None if conversion fails
                    df_expanded[EXPIRATION_DATE] = None

            if df_expanded.empty:
                return jsonify({ERROR_KEY: 'No valid domain information available to download'}), 404

            # Generate unique filename and save CSV
            base_csv_name = 'rk.csv'
            csv_filename = generate_unique_filename(base_csv_name)
            temp_dir = TEMP_FILE_PATH

            # Ensure temp directory exists
            os.makedirs(temp_dir, exist_ok=TRUE_VALUE)
            csv_path = os.path.join(temp_dir, csv_filename)
            df_expanded.to_csv(csv_path, index=FALSE_VALUE)

            current_dir = os.getcwd()
            return jsonify({
                MESSAGE_KEY: 'URL data processed successfully',
                'has_downloadable_data': TRUE_VALUE,
                PROGRESS_INFO_OBJ: df_expanded.to_dict('records'),
                'csv_filename': csv_filename,
                'current_directory': current_dir
            }), 200

        except Exception as processing_error:
            logging.error(f"Error processing results: {processing_error}")
            return jsonify({ERROR_KEY: f'Error processing results: {str(processing_error)}'}), 500

    except Exception as error:
        logging.error(f"Unexpected error in process_url_data: {error}")
        return jsonify({ERROR_KEY: f'Unexpected error: {str(error)}'}), 500
def join_array_values(arr):
    return ', '.join(map(STRING_TYPE, arr))

@app.route('/progress', methods=[POST_METHOD])
async def progress():
    cancel_flag = request.json.get(CANCEL_FLAG)
    if cancel_flag:
        print('cancelFlag At Progress Endpoint: ', cancel_flag)
        async for _ in process_urls_async({}, {}, cancel_flag):
            pass
        return jsonify(progress_info)
    return jsonify(progress_info)
@app.route('/download/<csvFilename>', methods=[POST_METHOD])
def download_csv(csvFilename):
    filename = csvFilename
    try:
        temp_dir = TEMP_FILE_PATH
        file_path = os.path.join(temp_dir, filename)
        print('File Path with File name: ', file_path)
        app.logger.info(f"File Path with File name: {file_path}")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as file:
                response = make_response(file.read())
            response.headers.set('Content-Type', 'text/csv')
            response.headers.set('Content-Disposition', f"attachment; filename={filename}")
            return response
        else:
            return jsonify({ERROR_KEY: 'File not found'})
    except Exception as error:
        return jsonify({ERROR_KEY: str(error)})
def list_directory_contents(current_directory, home_directory):
    directory = current_directory
    try:
        contents = os.listdir(directory)
        print('Current directory:', directory)
        print('Home directory:', home_directory)
        for item in contents:
            print(item)
    except OSError as error:
        print(f"Error: {error}")

HOME_DIR = os.path.expanduser('~')
CURRENT_DIR = os.getcwd()
# list_directory_contents(CURRENT_DIR, HOME_DIR)  # Commented out for debugging

def debug_form_data():
    file_type = request.form.get(FILE_KEY)
    batch = request.form.get('batch')
    url_column = request.form.get('url-column')
    sheet_name = request.form.get('sheet-name')
    column_name = request.form.get('column-name')
    required_data = request.form.get('required-data')
    output_file_type = request.form.get('output-file-type')
    print(f"File type: {file_type}, Batch: {batch}, URL column: {url_column}, Selected Sheet: {sheet_name}, Selected Column: {column_name}, Required data: {required_data}, Output file type: {output_file_type}")
    return 'Form submitted successfully'
# Import additional libraries needed for async operations
import aiohttp
import asyncio
import datetime
import whois
import time

# if __name__ == '__main__':
#     print("Starting URL Status Checker Flask Application...")
#     print("Visit http://localhost:5000 to access the application")
#     app.run(debug=True, host='0.0.0.0', port=5000)
